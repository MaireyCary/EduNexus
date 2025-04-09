# 处理Excel文件程序
import os
import zipfile
from . import parameter
import pandas as pd
from openpyxl import load_workbook


def class_get(file_path, sheet_name=0):
    """
    读取公共课表时间信息
    """
    # 读取 Excel 文件
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # 定义课表目标区域
    start_row = 2
    end_row = 6
    start_col = 2
    end_col = 6

    # 初始化字典来存储被占用的时间信息
    occupied_times = {}

    # 遍历目标区域
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            if pd.notnull(df.iloc[row, col]):  # 检查单元格是否被占用
                # 获取对应的时间信息
                time_info = parameter.time_mapping.get((row, col), None)
                if time_info:
                    day, period = time_info
                    if day not in occupied_times:
                        occupied_times[day] = set()
                    occupied_times[day].add(period)

    return occupied_times


def get_teacher_data(file_path, sheet_name=0):
    """
    读取教师表信息
    """
    # 读取 Excel 文件
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # 找到实际的表头行
    header_row = None
    for i in range(len(df)):
        if "班级名称" in df.iloc[i].values:
            header_row = i
            break

    if header_row is None:
        raise ValueError("无法找到表头行，请检查文件格式是否正确。")

    # 设置表头行
    df.columns = df.iloc[header_row]
    df = df[header_row + 1:]

    # 定义需要的列名
    columns = {
        "班级名称": "班级名称",
        "课程名称": "课程名称",
        "起止周": "起止周",
        "场地要求": "场地要求",
        "授课教师": "授课教师",
        "学分": "学分"
    }

    # 检查列名是否在实际列名中
    missing_columns = [col for col in columns.values() if col not in df.columns]
    if missing_columns:
        raise ValueError(f"以下列名在文件中不存在: {missing_columns}")

    # 筛选需要的列
    df = df[list(columns.values())]

    # 初始化结果字典
    result = {}

    # 遍历数据，按班级分类
    for _, row in df.iterrows():
        class_name = row["班级名称"]
        course_name = row["课程名称"]
        start_end_week = row["起止周"]
        venue_requirement = row["场地要求"]
        teacher = row["授课教师"]
        credit = row["学分"]

        # 如果班级名称为空，跳过当前行
        if pd.isna(class_name):
            continue

        # 如果班级名称尚未出现在结果字典中，初始化该班级
        if class_name not in result:
            result[class_name] = {}

        # 如果授课教师尚未出现在班级字典中，初始化该教师
        if teacher not in result[class_name]:
            result[class_name][teacher] = {}

        # 如果课程名称尚未出现在课程字典中，初始化该课程
        if course_name not in result[class_name][teacher]:
            result[class_name][teacher][course_name] = {
                "场地要求": venue_requirement
            }

        # 处理起止周部分
        if isinstance(start_end_week, str):
            # 如果起止周是字符串，则直接录入
            result[class_name][teacher][course_name]["起止周"] = start_end_week
        else:
            # 如果为空或者其他类型默认为1-18周
            result[class_name][teacher][course_name]["起止周"] = '1-18'

        # 处理学分部分
        if isinstance(credit, str):
            # 如果学分是字符串，取第一个字符并转换为整数
            result[class_name][teacher][course_name]["学分"] = int(credit[0])
        elif isinstance(credit, (int, float)):
            # 如果学分已经是数字类型，直接使用
            result[class_name][teacher][course_name]["学分"] = int(credit)
        else:
            # 如果学分是其他类型，设置为默认值 0
            result[class_name][teacher][course_name]["学分"] = 0

    return result


def write_excel(file_path, best_schedule):
    """
    将排课结果写入对应的课表中。
    """
    # 加载现有的课表
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 遍历排课结果，将其写入课表
    for (teacher, course), info in best_schedule.items():
        day = info["星期"]
        period = info["时间段"]
        place = info["教室"]
        week_range = info["起止周"]
        credit = info["学分"]

        # 获取时间映射
        for (row, col), (mapped_day, mapped_period) in parameter.time_mapping.items():
            if mapped_day == day and mapped_period == period:
                # 组合要写入的信息
                write_info = f"{course}/({period}){week_range}周/ {place}/{teacher}/{credit}"
                # 写入信息到对应的单元格
                sheet.cell(row=row + 1, column=col + 1).value = write_info
                break

    # 保存更新后的课表
    workbook.save(file_path)


def set_zip(files_path, zip_path):
    """
    将排课结果打包压缩，并存放到结果目录下。
    """
    if not os.path.exists(zip_path):
        os.makedirs(zip_path)

    # 定义压缩包的文件名
    zip_file_name = os.path.join(zip_path, "排课结果.zip")

    # 创建一个 ZipFile 对象
    with zipfile.ZipFile(zip_file_name, 'w', zipfile.ZIP_DEFLATED) as zip_f:
        # 遍历指定文件夹中的所有文件
        for root, dirs, files in os.walk(files_path):
            for file in files:
                # 获取文件的完整路径
                file_path = os.path.join(root, file)
                # 获取文件相对于 files_path 的相对路径
                rel_path = os.path.relpath(file_path, files_path)
                # 将文件添加到压缩包中
                zip_f.write(file_path, rel_path)


def get_zip(files_path):
    """
    解压上传上来的课表压缩包，并删除原压缩包。
    """
    # 确保文件夹路径存在
    if not os.path.exists(files_path):
        os.makedirs(files_path)

    # 遍历文件夹中的所有文件
    for file_name in os.listdir(files_path):
        file_path = os.path.join(files_path, file_name)

        # 检查文件是否是压缩包
        if zipfile.is_zipfile(file_path):
            # 创建一个 ZipFile 对象
            with zipfile.ZipFile(file_path, 'r', metadata_encoding='gbk') as zip_ref:
                # 解压所有文件到指定目录
                zip_ref.extractall(files_path)

            # 删除原压缩包
            os.remove(file_path)
